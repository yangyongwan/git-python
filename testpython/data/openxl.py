import openpyxl

class CasesData:
    '''用于保存测试用例数据'''
    pass
class Create:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name


    def create(self):
        wk = openpyxl.load_workbook()
        sh = wk.save()



    pass


class ReadExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def save(self):
        self.wk = openpyxl.Workbook()
        self.sh = self.wk.create_sheet(self.sheet_name)
        # self.sh.cell(row)
        self.wk.save(self.file_name)


    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sh = self.wb[self.sheet_name]

    def close(self):
        self.wb.close()

    def read_data(self):
        self.open()
        rows = list(self.sh.rows)
        titles = []
        for t in rows[0]:
            title = t.value
            titles.append(title)
        cases = []
        for row in rows[1:]:
            case = []
            for r in row:
                case.append(r.value)
            cases.append(dict(zip(titles,case)))
        self.close()
        return cases


    def write_data(self,row,colunn,msg):
        self.open()
        self.sh.cell(row=row,column=colunn,value=msg)
        self.wb.save(self.file_name)
        self.close()





if __name__ == '__main__':
    xh = ReadExcel(file_name='test1.xlsx',sheet_name='sheet1')
    # xh.save()
    xh.write_data(row=1,colunn=1,msg='账号')
    xh.write_data(row=1,colunn=2,msg='密码')
    xh.write_data(row=1,colunn=3,msg='信息')
    xh.write_data(row=2,colunn=1,msg='admin')
    xh.write_data(row=2,colunn=2,msg='Xh@2022!@#')
    xh.write_data(row=2,colunn=3,msg='现代化医院制度数字化应用平台')
    print(xh.read_data())
































#
# import openpyxl
#
# # 创建一个工作簿
# workbook = openpyxl.Workbook()
# # 创建一个表单
# sheet = workbook.create_sheet('表单1')
# # 写入一个数据
# sheet.cell(row=1, column=1, value="python")
# # 保存

#
# workbook = openpyxl.Workbook()
# sheet = workbook.create_sheet('表单')
# sheet.cell(row=1,column=1,value='python')
# workbook.save('test.xlsx')




#
# # 打开工作簿
# workbook = openpyxl.load_workbook('test.xlsx')
# # 获取表单
# sheet = workbook['表单1']
# # 读取指定的单元格数据
# cell = sheet.cell(row=1, column=1).value
# print







# # 方式一：读取A6单元格的值
# cell1 = sheet['A6'].value
#
# # 方式二：读取第3行,第4列单元格的值
# cell2 = sheet.cell(row=3, column=4).value
# # 读取A1-B4的单元格，共8个单元格
# cell3 = sheet['A1':'B4']
#
# # 读取A1-B4的单元格，共8个单元格
# cell4 = sheet['A1:B4']
#
# # 读取第2行的单元格
# cell5 = sheet[2]
#
# # 读取第1-2行的单元格
# cell5 = sheet[1:2]






